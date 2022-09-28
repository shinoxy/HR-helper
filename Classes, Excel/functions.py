from classes import *
from pathlib import Path
import openpyxl as o
from openpyxl import Workbook
import os
import re


def create_output(list_of_candidates, default_path):
    if os.path.isfile("Selected employees.xlsx"):  # Delete output file if exists
        os.remove("Selected employees.xlsx")
    selected_data = Workbook()  # Write selected data to output file
    output_data = selected_data.create_sheet(title='Selected employees', index=0)
    output_data['A1'] = 'Name'
    output_data['B1'] = 'Age'
    output_data['C1'] = 'Skills'

    for i in range(0, len(list_of_candidates)):
        output_data[i + 2][0].value = list_of_candidates[i].name
        output_data[i + 2][1].value = list_of_candidates[i].age
        output_data[i + 2][2].value = ', '.join(list_of_candidates[i].skills)

    save_data(selected_data, default_path)


def save_data(selected_data, default_path):
    print("Input directory or press any key to save results on Desktop by default")
    directory = input()
    try:
        selected_data.save(Path(directory, "Selected employees.xlsx"))
    except FileNotFoundError:
        selected_data.save(Path(default_path, "Selected employees.xlsx"))


def mode_1(list_of_candidates, requirements_list, default_path):
    print("Input directory of data:")
    try:
        input_data = o.open(Path(input()))  # Read data from input file
        data = input_data.active
        for row in range(2, data.max_row + 1):
            person = Person(data[row][0].value, data[row][1].value)
            person.skills = (re.sub(r'[^A-Za-z ]', '', data[row][2].value.lower())).split()
            list_of_candidates.append(person)
            for person in list_of_candidates:
                person.is_right(requirements_list)
                if person.status == "Rejected." or person.age < 18:
                    list_of_candidates.remove(person)

        create_output(list_of_candidates, default_path)

    except FileNotFoundError:
        print(f"No such file or directory")


def mode_2(list_of_candidates, requirements_list):
    while True:
        print("Input name")
        name = input()
        print("Input age")
        age = (int)(input())
        person = Person(name, age)
        if person.age >= 18:
            person.get_skills()
            person.is_right(requirements_list)
            if person.status == "Accepted!":
                list_of_candidates.append(person)
            person.change_skills()
        else:
            print("Forbidden")
        print("Print 'Y' if you want to add another person or press enter")
        if input().upper() != "Y":
            break

    for person in list_of_candidates:
        person.print_person_info()
